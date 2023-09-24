import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

# 1st Task = how many products we have per supplier 
# so first we have to check how many row are there in product list 

print(product_list.max_row)  

# create dictionaries as per task
products_per_supplier = {}
total_value_per_supplier = {}
Value_less_than_ten = {}

# 1st row only contains title : so we want it to operate from 2 
# range starts from 0 and not 1 

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    Product_num = product_list.cell(product_row, 1).value

#Calculation for number of products per supplier : 

    if supplier_name in products_per_supplier:
        current_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_products + 1

    else:
        print("adding a new supplier")
        products_per_supplier[supplier_name] = 1

#  calculation of total value per supplier : 
    
    if supplier_name in total_value_per_supplier:
        current_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_value + inventory * price

    else:
        print("adding a new supplier")
        total_value_per_supplier[supplier_name] = inventory * price

    # calculation for per supplier whose value is less than 10 :
    if inventory < 10:
        product_num = int(Product_num)
        
        Value_less_than_ten[product_num] = int(inventory)


print(products_per_supplier)
print(total_value_per_supplier)
print(Value_less_than_ten)


    